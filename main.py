import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# Set error tolerance range
tolerance = 0.0030

# Load data
normal_df = pd.read_excel('data/normal.xlsx')
cancer_df = pd.read_excel('data/cancer.xlsx')

# Ensure m/z data is of float type
normal_mz = normal_df.iloc[:, 0].astype(float)
cancer_mz = cancer_df.iloc[:, 0].astype(float)

# Get column names with prefixes to avoid duplicates
normal_cols = [f"Normal_{col}" for col in normal_df.columns]
cancer_cols = [f"Cancer_{col}" for col in cancer_df.columns]

# Rename columns to avoid conflicts
normal_df.columns = normal_cols
cancer_df.columns = cancer_cols

# Initialize result lists and matched indices
matched_rows = []
remaining_normal = []
matched_cancer_indices = set()

# Find matched items
for i, mz_normal in normal_mz.items():
    found_match = False
    for j, mz_cancer in cancer_mz.items():
        if abs(mz_normal - mz_cancer) <= tolerance and j not in matched_cancer_indices:
            combined_row = pd.concat([normal_df.iloc[i], cancer_df.iloc[j]], axis=0)
            matched_rows.append(combined_row)
            matched_cancer_indices.add(j)
            found_match = True
            break
    if not found_match:
        remaining_normal.append(normal_df.iloc[i])

# Find remaining cancer items
remaining_cancer = [cancer_df.iloc[j] for j in range(len(cancer_mz)) if j not in matched_cancer_indices]

# Create DataFrame for matched results
if matched_rows:
    matched_df = pd.DataFrame(matched_rows)
else:
    matched_df = pd.DataFrame(columns=normal_cols + cancer_cols)

# Write results to files
with pd.ExcelWriter('data/Common.xlsx', engine='openpyxl') as writer:
    matched_df.to_excel(writer, index=False, sheet_name='Common')

with pd.ExcelWriter('data/Normal only.xlsx', engine='openpyxl') as writer:
    pd.DataFrame(remaining_normal).to_excel(writer, index=False, sheet_name='Normal only', header=normal_cols)

with pd.ExcelWriter('data/Cancer only.xlsx', engine='openpyxl') as writer:
    pd.DataFrame(remaining_cancer).to_excel(writer, index=False, sheet_name='Cancer only', header=cancer_cols)

# Adjust formatting for matched file
if not matched_df.empty:
    wb = load_workbook('data/Common.xlsx')
    ws = wb['Common']
    for row in range(2, ws.max_row + 1):
        if ws[f'A{row}'].value == ws[f'A{row - 1}'].value:
            ws.merge_cells(f'A{row - 1}:A{row}')
            ws[f'A{row - 1}'].alignment = Alignment(horizontal="center", vertical="center")
    wb.save('data/Common.xlsx')

print("Comparison complete. Files generated: Common.xlsx, Normal only.xlsx, Cancer only.xlsx.")